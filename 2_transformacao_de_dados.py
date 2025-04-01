import zipfile
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

# Função para extrair o PDF do ZIP
def extract_pdf_from_zip(zip_path, pdf_filename):
    with zipfile.ZipFile(zip_path, 'r') as zipf:
        zipf.extract(pdf_filename, 'temp')  # Extraímos o PDF para a pasta 'temp'
    return os.path.join("temp", pdf_filename)

# Função para extrair dados estruturados do PDF
def extract_table_from_pdf(pdf_path):
    table_data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            print(f"Processando página {page_num}...")
            tables = page.extract_tables()

            if tables:
                for table in tables:
                    for row in table:
                        # Remove colunas vazias e espaços extras
                        cleaned_row = [cell.strip() if cell else "" for cell in row]
                        table_data.append(cleaned_row)
    
    return table_data

# Função para criar um Excel formatado
def save_to_excel(data, excel_filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rol de Procedimentos"

    # Estilização do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_style = Side(border_style="thin", color="000000")

    # Criar cabeçalhos baseados na primeira linha da tabela extraída
    if data:
        for col_num, value in enumerate(data[0], 1):
            cell = ws.cell(row=1, column=col_num, value=value)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # Adicionar os dados restantes
    for row_num, row in enumerate(data[1:], 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

       # Ajustar a largura das colunas automaticamente
     
    for col in ws.columns:
        column_letter = col[0].column_letter
        if column_letter in ["D", "E", "F", "H", "I", "J"]:  
            ws.column_dimensions[column_letter].width = 25  
        elif column_letter in ["L", "M"]:
            ws.column_dimensions[column_letter].width = 100  
        else:
            max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[column_letter].width = max_length + 2




    wb.save(excel_filename)
    print(f"Arquivo Excel '{excel_filename}' criado com sucesso!")

# Substituir abreviações por descrições completas
def replace_abbreviations_in_excel(excel_filename):
    abbreviations = {
        "OD": "Seg. Odontológica",
        "AMB": "Seg. Ambulatorial",
    }

    wb = openpyxl.load_workbook(excel_filename)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):  # Começar da linha 2 (pulando o cabeçalho)
        for cell in row:
            if cell.value in abbreviations:
                cell.value = abbreviations[cell.value]

    wb.save(excel_filename)
    print(f"As abreviações foram substituídas no arquivo {excel_filename}.")

# Função para compactar o arquivo Excel em ZIP
def zip_files(zip_name, files):
    with zipfile.ZipFile(zip_name, 'w') as zipf:
        for file in files:
            zipf.write(file, os.path.basename(file))  # Adiciona os arquivos com o nome base
    print(f"Arquivo ZIP '{zip_name}' criado com sucesso!")

# Caminho do arquivo ZIP que contém o PDF
zip_path = "Anexos.zip"
pdf_filename = "Anexo_I_Rol_2021RN_465.2021_RN627L.2024.pdf"

# Passo 1: Extrair o PDF do ZIP
pdf_path = extract_pdf_from_zip(zip_path, pdf_filename)

# Passo 2: Extrair dados da tabela
table_data = extract_table_from_pdf(pdf_path)

# Passo 3: Salvar os dados em um arquivo Excel
if table_data:
    excel_filename = "Rol_de_Procedimentos.xlsx"
    save_to_excel(table_data, excel_filename)

    # Passo 4: Substituir abreviações no Excel
    replace_abbreviations_in_excel(excel_filename)

    # Passo 5: Compactar o arquivo Excel em ZIP
    zip_filename = "Teste_Fernando_Caetano_de_Lima.zip"
    zip_files(zip_filename, [excel_filename])

    # Remover arquivos temporários
    os.remove(excel_filename)
    os.remove(pdf_path)

    print("Processo concluído com sucesso!")
else:
    print("Erro: Nenhuma tabela extraída do PDF.")
